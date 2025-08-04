"""
Obsługa zapisu i odczytu danych symulacji.

---
Simulation data saving and loading utilities.
"""
import json, datetime, os
import xml.etree.ElementTree as ET
import xml.dom.minidom
from pathlib import Path
from openpyxl import Workbook, load_workbook
from PyQt6.QtWidgets import QFileDialog, QMessageBox
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from tkinter import filedialog, messagebox, Listbox, MULTIPLE
import tkinter as tk

REQUIRED_USER_FIELDS = {
    "material", "T0", "Tc", "mu", "fluence",
    "pulse_duration", "laser_wavelength", "N", "asf"
}

ALLOWED_MATERIALS = ["Ni"]


def generate_graph():
    def select_file():
        filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if not filepath:
            return

        try:
            excel = pd.ExcelFile(filepath)
            sheet_names = excel.sheet_names

            listbox.delete(0, tk.END)
            for name in sheet_names:
                listbox.insert(tk.END, name)

            btn_generate.config(state="normal")
            root.filepath = filepath
            root.sheet_names = sheet_names
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load file: {e}")

    def generate_plots():
        selected_indices = listbox.curselection()
        if not selected_indices:
            messagebox.showwarning("Warning", "Please select at least one sheet.")
            return

        filepath = root.filepath
        selected_sheets = [root.sheet_names[i] for i in selected_indices]
        dataframes = {}

        try:
            for sheet in selected_sheets:
                df = pd.read_excel(filepath, sheet_name=sheet, skiprows=4)
                dataframes[sheet] = df
        except Exception as e:
            messagebox.showerror("Error", f"Error while reading sheets: {e}")
            return

        try:
            colors = plt.get_cmap("tab10")

            wykresy = [
                (1, "M3TM Koopmans electrons", "M3TM_Koopmans_electrons.png"),
                (4, "M3TM Koopmans phonons", "M3TM_Koopmans_phonons.png"),
                (7, "Magnetization", "graph_magnetization.png")
            ]

            for idx, (col_idx, tytul, filename) in enumerate(wykresy):
                plt.figure(figsize=(10, 6))
                rysowano = False

                for i, (sheet, df) in enumerate(dataframes.items()):
                    if df.shape[1] > col_idx:
                        x = df.iloc[:, 0]
                        y = df.iloc[:, col_idx]

                        if pd.api.types.is_numeric_dtype(y):
                            plt.plot(x, y, label=sheet, color=colors(i))
                            rysowano = True

                if rysowano:
                    plt.title(tytul)
                    plt.xlabel("czas")
                    plt.ylabel(tytul)
                    plt.legend()
                    plt.grid(True)
                    plt.tight_layout()

                    out_path = os.path.join(os.path.dirname(filepath), filename)
                    plt.savefig(out_path)
                plt.close()

            messagebox.showinfo("Success", "Plots have been saved.")
        except Exception as e:
            messagebox.showerror("Error", f"Could not generate plots: {e}")

    root = tk.Tk()
    root.title("Excel Plot Generator")

    btn_file = tk.Button(root, text="Select Excel file", command=select_file)
    btn_file.pack(pady=10)

    tk.Label(root, text="Select sheets:").pack()
    listbox = Listbox(root, selectmode=MULTIPLE, width=50)
    listbox.pack(pady=5)

    btn_generate = tk.Button(root, text="Generate plots", command=generate_plots, state="disabled")
    btn_generate.pack(pady=10)

    root.mainloop()

def save_simulation_to_excel(params, plot_data):
    """
       Zapisuje dane symulacji do pliku Excel (XLSX) w katalogu uruchomienia aplikacji.

       Funkcja tworzy (jeśli nie istnieje) lub otwiera plik o nazwie 'Simulations.xlsx' w folderze,
       w którym uruchamiana jest aplikacja. Dla każdej symulacji tworzony jest nowy arkusz,
       którego nazwa zawiera liczbę warstw i moc lasera, np.: 'N=4, Fluence=2.5'.

       Jeśli arkusz o danej nazwie już istnieje, funkcja doda numer sufiksu (np. _1, _2 itd.),
       aby uniknąć nadpisywania danych.

       Parametry wejściowe zapisywane są w pierwszych dwóch wierszach arkusza,
       a dane z wykresów liniowych w układzie poziomym (każdy wykres w dwóch kolumnach: x i y),
       z jedną pustą kolumną odstępu między kolejnymi wykresami.

       Args:
           params (dict): Parametry symulacji do zapisania. Muszą zawierać co najmniej:
                          - "liczba_warstw" (int)
                          - "moc_lasera" (float lub str)
           plot_data (dict): Dane wykresów w formacie {"lines": [{"title": ..., "x": [...], "y": [...]}]}.

       Returns:
           None

       Raises:
           Exception: Gdy zapis do pliku się nie powiedzie.

       ---
       Saves simulation data to an Excel (XLSX) file in the application's execution directory.

       The function creates (if missing) or opens 'Simulations.xlsx' and adds a new sheet for each simulation.
       The sheet name is based on the number of layers and laser fluence, e.g., 'N=4, Fluence=2.5'.

       If a sheet with that name already exists, the function adds a numeric suffix to avoid overwriting.

       The function writes input parameters in the first two rows,
       and line chart data in a horizontal layout (each chart gets two columns: x and y),
       with one empty column of spacing between charts.

       Args:
           params (dict): Simulation parameters to save. Must include at least:
                          - "liczba_warstw" (int)
                          - "moc_lasera" (float or str)
           plot_data (dict): Plot data in the format {"lines": [{"title": ..., "x": [...], "y": [...]}]}.

       Returns:
           None

       Raises:
           Exception: If saving the file fails.
       """
    current_dir = Path.cwd()
    filename = current_dir / "Simulations.xlsx"

    num_layers = params.get("N", "unknown")
    fluence = params.get("fluence", "unknown")
    sheet_name = f"N={num_layers}, Fluence={fluence}"

    if filename.exists():
        wb = load_workbook(filename)
    else:
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

    original_name = sheet_name
    counter = 1
    while sheet_name in wb.sheetnames:
        sheet_name = f"{original_name}_{counter}"
        counter += 1

    ws = wb.create_sheet(title=sheet_name)

    parameter_keys = list(params.keys())
    ws.append(parameter_keys)
    parameter_values = [params[key] for key in parameter_keys]
    ws.append(parameter_values)
    ws.append([])

    if 'lines' in plot_data:
        max_len = 0
        for line in plot_data['lines']:
            max_len = max(max_len, len(line.get('x', [])))

        col = 1
        for i, line in enumerate(plot_data['lines']):
            x_vals = line.get('x', [])
            y_vals = line.get('y', [])
            title = line.get('title', f"Line {i + 1}")

            ws.cell(row=4, column=col, value=f"{title} - x")
            ws.cell(row=4, column=col + 1, value=f"{title} - y")

            for row_i in range(max_len):
                x = x_vals[row_i] if row_i < len(x_vals) else None
                y = y_vals[row_i] if row_i < len(y_vals) else None
                ws.cell(row=5 + row_i, column=col, value=x)
                ws.cell(row=5 + row_i, column=col + 1, value=y)

            col += 3

    wb.save(filename)
    print(f"File saved to: {filename}, sheet: {sheet_name}")

def save_simulation_parameters(params, file_path, file_format, parameter_encoder, quantity_to_plain_func):
    """
    Zapisuje dane symulacji w formacie JSON lub XML.

    Args:
        params (dict): Parametry symulacji do zapisania.
        file_path (str): Ścieżka pliku docelowego (bez rozszerzenia).
        file_format (str): Format zapisu ("json" lub "xml").
        parameter_encoder (type): Niestandardowy encoder JSON.
        quantity_to_plain_func (callable): Funkcja konwertująca jednostki na wartości tekstowe.

    Returns:
        str: Pełna ścieżka zapisanego pliku.

    Raises:
        Exception: Gdy zapis się nie powiedzie.

    ---
    Saves simulation parameters to JSON or XML format.

    Args:
        params (dict): Parameters to be saved.
        file_path (str): Output file path (without extension).
        file_format (str): File format ("json" or "xml").
        parameter_encoder (type): Custom JSON encoder class.
        quantity_to_plain_func (callable): Function to convert quantities to plain values.

    Returns:
        str: Full path to the saved file.

    Raises:
        Exception: If saving fails.
    """
    if file_format == "json":
        if not file_path.lower().endswith(".json"):
            file_path += ".json"
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(params, f, indent=2, ensure_ascii=False, cls=parameter_encoder)
        return file_path

    elif file_format == "xml":
        if not file_path.lower().endswith(".xml"):
            file_path += ".xml"

        plain_params = quantity_to_plain_func(params)
        root = ET.Element("simulation_parameters")

        def add_dict_to_xml(parent, data):
            if isinstance(data, dict):
                for key, value in data.items():
                    child = ET.SubElement(parent, str(key))
                    add_dict_to_xml(child, value)
            elif isinstance(data, list):
                for item in data:
                    item_elem = ET.SubElement(parent, "item")
                    add_dict_to_xml(item_elem, item)
            else:
                parent.text = str(data)

        add_dict_to_xml(root, plain_params)

        rough_string = ET.tostring(root, 'utf-8')
        reparsed = xml.dom.minidom.parseString(rough_string)
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(reparsed.toprettyxml(indent="  "))

        return file_path

    else:
        raise ValueError("Unsupported file format: " + file_format)

def load_simulation_parameters(file_path,parent_widget):
    """
       Wczytuje dane symulacji z pliku JSON.

       Args:
           file_path (str): Ścieżka do pliku .json
           parent_widget (QWidget): Rodzic do wyświetlenia komunikatów błędów.

       Returns:
           dict: Dane do wypełnienia formularza symulacji.

       Raises:
           ValueError: Gdy format pliku lub dane są niepoprawne.

       ---
       Loads simulation parameters from a JSON file.

       Args:
           file_path (str): Path to the .json file.
           parent_widget (QWidget): Parent widget for displaying error messages.

       Returns:
           dict: Parameters for simulation form filling.

       Raises:
           ValueError: If file or data is invalid.
       """
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        if not isinstance(data, dict):
            error_message_format = parent_widget.error_invalid_format.format(fields=", ".join(REQUIRED_USER_FIELDS))
            raise ValueError(error_message_format)

        missing = REQUIRED_USER_FIELDS - data.keys()
        if missing:
            error_message = parent_widget.missing_fields_error.format(fields=", ".join(missing))
            raise ValueError(error_message)

        for key in REQUIRED_USER_FIELDS:
            if key == "material":
                continue
            value = data[key]
            if not isinstance(value, (int, float)):
                raise ValueError(parent_widget.error_invalid_type.format(field=key))

        material = data.get("material")
        if material not in ALLOWED_MATERIALS:
            raise ValueError(parent_widget.error_invalid_material.format(
                material=material,
                allowed=", ".join(ALLOWED_MATERIALS)
            ))

        return {key: data[key] for key in REQUIRED_USER_FIELDS}

    except json.JSONDecodeError:
        raise ValueError(parent_widget.error_json_decode)

    except FileNotFoundError:
        raise ValueError(parent_widget.error_file_not_found.format(path=file_path))

def save_simulation_report(params, material_name, material_props, plot_data, parent=None, simulation_duration=None):
    """
    Tworzy i zapisuje raport z przebiegu symulacji do pliku tekstowego.

    Args:
        params (dict): Parametry wejściowe symulacji.
        material_name (str): Nazwa materiału.
        material_props (dict): Właściwości materiału.
        plot_data (dict): Dane do wykresów (linie, mapy).
        parent (QWidget, optional): Rodzic do komunikatów GUI.
        simulation_duration (float, optional): Czas trwania symulacji w sekundach.

    Returns:
        None

    ---
    Generates and saves a simulation report to a text file.

    Args:
        params (dict): Input parameters of the simulation.
        material_name (str): Name of the material.
        material_props (dict): Material properties.
        plot_data (dict): Plot data (lines and maps).
        parent (QWidget, optional): Parent widget for error messages.
        simulation_duration (float, optional): Simulation duration in seconds.

    Returns:
        None
    """
    file_path, selected_filter = QFileDialog.getSaveFileName(
        parent,
        "Save simulation report",
        "",
        "Text files (*.txt);;All files (*)",
        options=QFileDialog.Option.DontUseNativeDialog
    )

    if not file_path:
        return

    if not os.path.splitext(file_path)[1]:
        file_path += ".txt"

    try:
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write("LaserDeMag simulation report\n")
            f.write(f"Simulation date: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            if simulation_duration is not None:
                f.write(f"Simulation duration: {simulation_duration:.2f} seconds\n")
            f.write("\n--- Simulation parameters ---\n")
            for key, val in params.items():
                f.write(f"{key}: {val}\n")

            f.write(f"\n--- Material ---\n")
            f.write(f"Name: {material_name}\n")
            for prop_key, prop_val in material_props.items():
                f.write(f"{prop_key}: {prop_val}\n")

            f.write("\n--- Simulation results (selected data) ---\n")
            if 'maps' in plot_data:
                f.write("Data maps:\n")

                maps = plot_data['maps']
                delays = maps.get("delays", [])
                distances = maps.get("distances", [])
                temp_map = maps.get("temp_map", None)

                f.write(f"  delays: {delays[:5]}... (total {len(delays)})\n")
                f.write(f"  distances: {distances[:5]}... (total {len(distances)})\n")

                if isinstance(temp_map, (list, np.ndarray)):
                    arr = np.array(temp_map)
                    f.write(f"  temp_map shape: {arr.shape} (time, space, component)\n")

                    if arr.ndim == 3 and arr.shape[2] > 2:
                        f.write("  Sample values (T_spin):\n")
                        for t in range(min(2, arr.shape[0])):
                            f.write(f"    {arr[t, :5, 2]}...\n")
                else:
                    f.write("  temp_map: missing or invalid format\n")

            if 'lines' in plot_data:
                f.write("Line plots:\n")
                for i, d in enumerate(plot_data['lines']):
                    f.write(f"  Line {i + 1} - {d.get('title', '')}:\n")
                    f.write(f"    x: {d.get('x', [])[:5]}... (total {len(d.get('x', []))})\n")
                    f.write(f"    y: {d.get('y', [])[:5]}... (total {len(d.get('y', []))})\n")

            f.write("\n--- End of report ---\n")

        QMessageBox.information(
            parent,
            "Save successful",
            f"The report was successfully saved to:\n{file_path}"
        )

    except Exception as e:
        QMessageBox.critical(
            parent,
            "Save error",
            f"An error occurred while saving the report:\n{str(e)}"
        )